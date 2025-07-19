import openpyxl as vb
path=r'e:/python/project/epem.xlsx'
path2=r'e:/python/project/loadcal.xlsx'
#path=r'd:/python/epem.xlsx'
#path2=r'd:/python/loadcal.xlsx'
wb=vb.load_workbook(path, data_only=True)
wb2=vb.load_workbook(path2)
#for n in range(0,len(wb.sheetnames)):
for n in range(0,len(wb.sheetnames)):
    wb2.worksheets[n].title = wb.worksheets[n].title
    for i in range(5,60):
        wb2.worksheets[n].cell(i, 2).value = wb.worksheets[n].cell(i - 4, 2).value  #设备名称
        wb2.worksheets[n].cell(i, 3).value = wb.worksheets[n].cell(i - 4, 1).value  #工艺代号
        wb2.worksheets[n].cell(i, 5).value = wb.worksheets[n].cell(i - 4, 5).value  #设备功率
        wb2.worksheets[n].cell(i, 6).value = wb.worksheets[n].cell(i - 4, 3).value  #常用台数
wb2.save(path2)
