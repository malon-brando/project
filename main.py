import openpyxl as vb
path=r'e:/python/project/7L7-MCC.xlsx'
#path2=r'e:/python/project/epem.xlsx'
wb=vb.load_workbook(path)
for n in range(0,len(wb.sheetnames)):
    for i in range(1, wb.worksheets[n].max_row+1):
        aa1=wb.worksheets[n].cell(i, 4).value
        if aa1 is None:
            continue
        if len(aa1)<5:
            pass
        else:
            aa1=aa1[3:]
            wb.worksheets[n].cell(i, 4).value=aa1


'''for i in range(1,wb.worksheets[0].max_row+1):
    a=wb.worksheets[0].cell(i, 1).value
    wb.worksheets[0].cell(i, 1).value=a[0:3]+'.'+a[4:6]
    #print(wb.worksheets[0].cell(i, 1).value)
    b = wb.worksheets[0].cell(i, 2).value
    wb.worksheets[0].cell(i, 2).value = b[0:7]
'''
wb.save(path)
