import openpyxl as vb
path=r'd:/python/project/charm.xlsx'
path2=r'd:/python/project/epem.xlsx'
#path=r'd:/python/charm.xlsx'
#path2=r'd:/python/epem.xlsx'
wb=vb.load_workbook(path)
wb2=vb.Workbook(path2)
wb2.save(path2)
wb2=vb.load_workbook(path2)  #新建epem表，不用再受到创建
for n in range(0,len(wb.sheetnames)):
    sheetxx = wb2.create_sheet("n", n)  # 新建sheet表
    wb2.worksheets[n].title=wb.worksheets[n].title
    wb2.worksheets[n].cell(1, 5).value = '=D1/C1'
    kk=1
    maxline=wb.worksheets[n].max_row
    for i in range(6,wb.worksheets[n].max_row):       #原表从功率那一列开始遍历
        if wb.worksheets[n].cell(i,12).value is None:
            continue
        else:
           wb2.worksheets[n].cell(kk,4).value=wb.worksheets[n].cell(i,12).value #写入新表功率值
           wb2.worksheets[n].cell(kk,3).value=wb.worksheets[n].cell(i,10).value #写入新表,用电设备数量
           for k in range(i,5,-1):     #遍历原表车间号，k为从第I行往前推，得到的行数
               if wb.worksheets[n].cell(k,1).value is None:
                   continue
               else:                                      #写入新表车间号
                   wb2.worksheets[n].cell(kk, 1).value=str(wb.worksheets[n].cell(k, 1).value)+\
                       str(wb.worksheets[n].cell(k,2).value)+str(wb.worksheets[n].cell(i, 3).value)
                   #上方是写入车间号，314,是吧第一列和第二列加一块，目前看都是在一块
               for sk in range(i,5,-1):               #写入第二列的名称
                   if wb.worksheets[n].cell(sk, 2).value==wb.worksheets[n].cell(k, 2).value:
                       aoe=sk
                       continue
               break
           wb2.worksheets[n].cell(kk, 2).value = str(wb.worksheets[n].cell(aoe, 4).value) + ' ' + \
                                                 str(wb.worksheets[n].cell(i, 4).value)
           kk += 1
wb2.save(path2)